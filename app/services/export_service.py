"""
导出任务服务
"""
import os
import logging
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.config import BASE_DIR
from app.database import SessionLocal
from app.models.export_task import ExportTask
from app.models.zabbix_config import ZabbixConfig
from app.models.monitor_filter import MonitorFilter
from app.services.zabbix_service import ZabbixService
from app.services.monitor_service import MonitorService

logger = logging.getLogger(__name__)

# 导出目录
EXPORT_DIR = os.path.join(BASE_DIR, "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


def do_export(task_id: int):
    """执行导出任务"""
    db = SessionLocal()

    try:
        task = db.query(ExportTask).filter(ExportTask.id == task_id).first()
        if not task:
            logger.error(f"导出任务不存在: {task_id}")
            return

        # 更新状态为处理中
        task.status = "processing"
        db.commit()

        # 获取Zabbix配置
        zabbix_config = db.query(ZabbixConfig).filter(
            ZabbixConfig.id == task.zabbix_config_id
        ).first()
        if not zabbix_config:
            raise Exception("Zabbix配置不存在")

        # 创建服务实例
        zabbix_service = ZabbixService.from_config(zabbix_config)
        monitor_service = MonitorService(zabbix_service)

        # 获取监控数据
        hosts_data = zabbix_service.get_all_hosts_with_status()
        monitor_data = monitor_service.get_multiple_filters_data(
            db, task.filter_ids, 86400  # 默认24小时
        )

        # 生成Excel文件
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        filename = f"{zabbix_config.name}_{timestamp}.xlsx"
        filepath = os.path.join(EXPORT_DIR, filename)

        _generate_excel(
            filepath=filepath,
            zabbix_config_name=zabbix_config.name,
            hosts_data=hosts_data,
            monitor_data=monitor_data,
            filter_ids=task.filter_ids,
            include_device_overview=task.include_device_overview,
            db=db
        )

        # 更新任务状态
        task.status = "completed"
        task.file_path = filepath
        task.filename = filename
        task.completed_at = datetime.now()
        db.commit()

        logger.info(f"导出任务完成: {task_id}")

    except Exception as e:
        logger.error(f"导出任务失败: {e}")
        if task:
            task.status = "failed"
            task.error_message = str(e)
            task.completed_at = datetime.now()
            db.commit()

    finally:
        db.close()


def _generate_excel(filepath: str, zabbix_config_name: str, hosts_data: dict,
                    monitor_data: dict, filter_ids: list,
                    include_device_overview: bool, db):
    """
    生成Excel报告
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 设备概览页
    if include_device_overview:
        ws = wb.create_sheet("设备概览")
        ws.merge_cells('A1:C1')
        cell = ws.cell(row=1, column=1, value="设备概览")
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=3, column=1, value="设备总量").font = Font(bold=True, size=14)
        ws.cell(row=3, column=2, value=hosts_data["total"])
        ws.cell(row=4, column=1, value="在线设备量").font = Font(bold=True, size=14)
        ws.cell(row=4, column=2, value=len(hosts_data["online"]))
        ws.cell(row=5, column=1, value="离线设备量").font = Font(bold=True, size=14)
        ws.cell(row=5, column=2, value=len(hosts_data["offline"]))

        # 离线设备列表
        if hosts_data["offline"]:
            ws.cell(row=7, column=1, value="离线设备列表").font = Font(bold=True, size=14)
            headers = ["设备名", "IP地址", "所属群组"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=8, column=col, value=header).font = Font(bold=True)
            for idx, host in enumerate(hosts_data["offline"], 9):
                ws.cell(row=idx, column=1, value=host["name"]).font = Font(color="FF0000")
                ws.cell(row=idx, column=2, value=host["ip"]).font = Font(color="FF0000")
                ws.cell(row=idx, column=3, value=host["groups"]).font = Font(color="FF0000")

    # 为每个筛选配置创建工作表
    for filter_id in filter_ids:
        filter_obj = db.query(MonitorFilter).filter(MonitorFilter.id == filter_id).first()
        if not filter_obj:
            continue

        filter_data = monitor_data.get(filter_obj.name, [])
        if not filter_data:
            continue

        # 工作表名称最长31字符
        sheet_name = filter_obj.name[:31]
        ws = wb.create_sheet(sheet_name)

        # 标题
        ws.merge_cells('A1:H1')
        cell = ws.cell(row=1, column=1, value=filter_obj.name)
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # 表头
        headers = ["所属群组", "设备名", "IP地址", "监控项", "当前值", "最大值", "最小值", "平均值"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header).font = Font(bold=True)

        # 数据
        for row_idx, row_data in enumerate(filter_data, 3):
            ws.cell(row=row_idx, column=1, value=row_data.get("groups", ""))
            ws.cell(row=row_idx, column=2, value=row_data.get("hostname", ""))
            ws.cell(row=row_idx, column=3, value=row_data.get("ip", ""))
            ws.cell(row=row_idx, column=4, value=row_data.get("item_name", ""))
            ws.cell(row=row_idx, column=5, value=row_data.get("current", 0))
            ws.cell(row=row_idx, column=6, value=row_data.get("max", 0))
            ws.cell(row=row_idx, column=7, value=row_data.get("min", 0))
            ws.cell(row=row_idx, column=8, value=row_data.get("avg", 0))

    wb.save(filepath)
