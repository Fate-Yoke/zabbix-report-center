"""
邮件服务
"""
import os
import smtplib
import logging
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.header import Header
from urllib.parse import quote
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.services.encryption_service import encryption_service

logger = logging.getLogger(__name__)


class EmailService:
    """邮件服务类"""

    def __init__(self, smtp_server: str, smtp_port: int, smtp_user: str,
                 smtp_pass: str, mail_from: str, use_ssl: bool = True):
        self.smtp_server = smtp_server
        self.smtp_port = smtp_port
        self.smtp_user = smtp_user
        self.smtp_pass = smtp_pass
        self.mail_from = mail_from
        self.use_ssl = use_ssl

    @classmethod
    def from_config(cls, config) -> "EmailService":
        """从配置对象创建实例"""
        smtp_pass = encryption_service.decrypt(config.smtp_pass)
        return cls(
            smtp_server=config.smtp_server,
            smtp_port=config.smtp_port,
            smtp_user=config.smtp_user,
            smtp_pass=smtp_pass,
            mail_from=config.mail_from,
            use_ssl=config.use_ssl
        )

    def test_connection(self) -> dict:
        """测试邮件连接"""
        server = None
        try:
            if self.use_ssl:
                server = smtplib.SMTP_SSL(self.smtp_server, self.smtp_port, timeout=20)
            else:
                server = smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=20)
                if server.has_extn("STARTTLS"):
                    server.starttls()

            server.login(self.smtp_user, self.smtp_pass)
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}
        finally:
            if server:
                try:
                    server.quit()
                except:
                    pass

    def send_email(self, to_addrs: List[str], subject: str,
                   body: str, attachment_path: str = None) -> dict:
        """
        发送邮件
        """
        msg = MIMEMultipart()
        msg["From"] = self.mail_from
        msg["To"] = ",".join(to_addrs)
        msg["Subject"] = subject

        msg.attach(MIMEText(body, "plain", "utf-8"))

        # 添加附件
        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, "rb") as f:
                part = MIMEBase(
                    "application",
                    "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                part.set_payload(f.read())
                encoders.encode_base64(part)

                # 处理中文文件名
                filename = os.path.basename(attachment_path)
                try:
                    # 尝试 ASCII 编码，如果失败则使用 RFC 2231 编码
                    filename.encode('ascii')
                    part.add_header(
                        "Content-Disposition",
                        f'attachment; filename="{filename}"'
                    )
                except UnicodeEncodeError:
                    # 文件名包含非 ASCII 字符，使用 RFC 2231 编码
                    encoded_filename = quote(filename, safe='')
                    part.add_header(
                        "Content-Disposition",
                        f"attachment; filename*=UTF-8''{encoded_filename}"
                    )
                msg.attach(part)

        server = None
        try:
            # 尝试SSL连接
            try:
                logger.info("尝试 SMTP_SSL 连接...")
                server = smtplib.SMTP_SSL(self.smtp_server, self.smtp_port, timeout=20)
                server.ehlo()
                logger.info("SMTP_SSL 连接成功")
            except Exception as e:
                logger.info(f"SMTP_SSL 不可用，改用 SMTP：{e}")
                server = smtplib.SMTP(self.smtp_server, self.smtp_port, timeout=20)
                server.ehlo()

            server.login(self.smtp_user, self.smtp_pass)
            server.send_message(msg)
            logger.info("邮件发送成功")

            return {"success": True}
        except Exception as e:
            logger.error(f"邮件发送失败: {e}")
            return {"success": False, "error": str(e)}
        finally:
            if server:
                try:
                    server.quit()
                except:
                    pass

    @staticmethod
    def generate_excel_report(data: dict, output_path: str) -> str:
        """
        生成Excel报告
        data: {
            "total": int,
            "online": list,
            "offline": list,
            "monitor_data": {filter_name: [...]}
        }
        """
        wb = Workbook()
        wb.remove(wb.active)

        # 设备概览页
        ws = wb.create_sheet("设备概览")
        ws.merge_cells('A1:C1')
        cell = ws.cell(row=1, column=1, value="设备概览")
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        rows = [
            ("设备总量", data.get("total", 0)),
            ("在线设备量", len(data.get("online", []))),
            ("离线设备量", len(data.get("offline", [])))
        ]
        for i, (title, value) in enumerate(rows, start=2):
            ws.cell(row=i, column=1, value=title).font = Font(bold=True, size=14)
            ws.cell(row=i, column=2, value=value)

        # 离线设备列表
        ws.cell(row=6, column=1, value="离线设备列表").font = Font(bold=True, size=14)
        ws.append(["设备名", "IP地址", "所属群组"])
        for col in range(1, 4):
            ws.cell(row=7, column=col).font = Font(bold=True)

        for idx, host in enumerate(data.get("offline", []), start=8):
            ws.append([host.get("name", ""), host.get("ip", ""), host.get("groups", "")])
            for c in range(1, 4):
                ws.cell(row=idx, column=c).font = Font(color="FF0000")

        # 各监控数据页
        for filter_name, filter_data in data.get("monitor_data", {}).items():
            ws_monitor = wb.create_sheet(filter_name[:31])  # Excel工作表名最多31字符
            ws_monitor.merge_cells('A1:H1')
            cell = ws_monitor.cell(row=1, column=1, value=filter_name)
            cell.font = Font(bold=True, size=16)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # 表头
            headers = ["所属群组", "设备名", "IP地址", "监控项", "当前值", "最大值", "最小值", "平均值"]
            ws_monitor.append(headers)
            for col in range(1, len(headers) + 1):
                ws_monitor.cell(row=2, column=col).font = Font(bold=True)

            # 数据
            for row_data in filter_data:
                ws_monitor.append([
                    row_data.get("groups", ""),
                    row_data.get("hostname", ""),
                    row_data.get("ip", ""),
                    row_data.get("item_name", ""),
                    row_data.get("current", 0),
                    row_data.get("max", 0),
                    row_data.get("min", 0),
                    row_data.get("avg", 0)
                ])

        wb.save(output_path)
        return output_path
