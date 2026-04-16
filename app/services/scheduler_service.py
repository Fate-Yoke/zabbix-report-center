"""
定时任务调度服务
"""
import os
import time
import logging
from datetime import datetime
from typing import Dict, Any

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from apscheduler.jobstores.memory import MemoryJobStore
from apscheduler.executors.pool import ThreadPoolExecutor
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.config import BASE_DIR
from app.models.schedule_task import ScheduleTask, TaskLog
from app.models.zabbix_config import ZabbixConfig
from app.models.email_config import EmailConfig
from app.models.monitor_filter import MonitorFilter
from app.services.zabbix_service import ZabbixService
from app.services.email_service import EmailService
from app.services.monitor_service import MonitorService
from app.database import SessionLocal

logger = logging.getLogger(__name__)

# 导出目录
EXPORT_DIR = os.path.join(BASE_DIR, "exports")
os.makedirs(EXPORT_DIR, exist_ok=True)


def execute_task(task_id: int):
    """执行任务 - 独立函数，避免序列化问题"""
    db = SessionLocal()
    log = None

    try:
        task = db.query(ScheduleTask).filter(ScheduleTask.id == task_id).first()
        if not task:
            logger.error(f"任务不存在: {task_id}")
            return

        # 创建日志记录
        log = TaskLog(
            task_id=task_id,
            status="running",
            started_at=datetime.now()
        )
        db.add(log)
        db.commit()

        # 获取Zabbix配置
        zabbix_config = db.query(ZabbixConfig).filter(
            ZabbixConfig.id == task.zabbix_config_id
        ).first()
        if not zabbix_config:
            raise Exception("Zabbix配置不存在")

        # 获取邮件配置
        email_config = None
        if task.email_config_id:
            email_config = db.query(EmailConfig).filter(
                EmailConfig.id == task.email_config_id
            ).first()

        # 创建服务实例
        zabbix_service = ZabbixService.from_config(zabbix_config)
        monitor_service = MonitorService(zabbix_service)

        # 获取监控数据
        hosts_data = zabbix_service.get_all_hosts_with_status()
        monitor_data = monitor_service.get_multiple_filters_data(
            db, task.monitor_filter_ids, task.time_range
        )

        # 生成Excel文件
        output_path, output_filename = _generate_excel(
            zabbix_config_name=zabbix_config.name,
            hosts_data=hosts_data,
            monitor_data=monitor_data,
            filter_ids=task.monitor_filter_ids,
            include_device_overview=task.include_device_overview,
            include_alert_data=task.include_alert_data,
            zabbix_service=zabbix_service,
            db=db
        )

        # 生成邮件标题
        subject_parts = []
        if task.email_subject:
            subject_parts.append(task.email_subject)

        if task.subject_suffix_config_name:
            subject_parts.append(zabbix_config.name)

        if task.subject_suffix_timestamp:
            timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
            subject_parts.append(timestamp)

        if not subject_parts:
            subject_parts.append(f"监控报告-{task.name}")

        subject = "_".join(subject_parts)

        # 生成邮件内容
        body_parts = []

        # 添加自定义邮件内容
        if task.email_body:
            body_parts.append(task.email_body)

        # 添加设备概览信息
        if task.email_include_device_overview:
            device_overview = f"""设备概览：
设备总量：{hosts_data['total']}
在线设备量：{len(hosts_data['online'])}
离线设备量：{len(hosts_data['offline'])}"""
            if hosts_data['offline']:
                device_overview += "\n\n离线设备列表："
                for host in hosts_data['offline'][:10]:  # 最多显示10个
                    device_overview += f"\n  - {host['name']} ({host['ip']})"
                if len(hosts_data['offline']) > 10:
                    device_overview += f"\n  ... 等 {len(hosts_data['offline'])} 台设备"
            body_parts.append(device_overview)

        # 添加监控数据摘要
        if task.email_include_monitor_summary and monitor_data:
            monitor_summary = "监控数据摘要："
            for filter_name, filter_data in monitor_data.items():
                if filter_data:
                    monitor_summary += f"\n  {filter_name}: {len(filter_data)} 条记录"
            body_parts.append(monitor_summary)

        # 添加告警数据摘要
        if task.email_include_alert_summary:
            try:
                # 获取最近24小时的告警数据
                time_till = int(time.time())
                time_from = time_till - 86400  # 24小时前
                alerts = zabbix_service.get_alerts(
                    time_from=time_from,
                    time_till=time_till
                )

                # 按级别统计告警
                alert_stats = {
                    "灾难": 0,
                    "严重": 0,
                    "一般严重": 0,
                    "警告": 0,
                    "信息": 0,
                    "未分类": 0
                }
                severity_map = {
                    5: "灾难",
                    4: "严重",
                    3: "一般严重",
                    2: "警告",
                    1: "信息",
                    0: "未分类"
                }

                for alert in alerts:
                    severity_name = severity_map.get(alert.get("severity", 0), "未分类")
                    alert_stats[severity_name] += 1

                alert_summary = f"告警数据摘要（最近24小时）：\n总告警数：{len(alerts)}"
                for severity_name, count in alert_stats.items():
                    if count > 0:
                        alert_summary += f"\n  {severity_name}: {count} 条"

                body_parts.append(alert_summary)
            except Exception as e:
                logger.warning(f"获取告警数据摘要失败: {e}")

        body_parts.append("\n详细数据请查看附件。")
        body = "\n\n".join(body_parts)

        # 发送邮件
        if email_config:
            email_service = EmailService.from_config(email_config)
            result = email_service.send_email(
                to_addrs=task.recipients,
                subject=subject,
                body=body,
                attachment_path=output_path
            )

            if result["success"]:
                log.status = "success"
                log.message = "邮件发送成功"
            else:
                log.status = "failed"
                log.message = f"邮件发送失败: {result.get('error', '未知错误')}"
        else:
            log.status = "success"
            log.message = "报告已生成(未配置邮件)"

        log.recipients = task.recipients
        log.attachment_path = output_path
        log.attachment_filename = output_filename
        log.finished_at = datetime.now()

        # 更新任务执行时间
        task.last_run = datetime.now()
        db.commit()

        logger.info(f"任务执行完成: {task.name}")

    except Exception as e:
        logger.error(f"任务执行失败: {e}")
        if log:
            log.status = "failed"
            log.message = str(e)
            log.finished_at = datetime.now()
            db.commit()

    finally:
        db.close()


def _generate_excel(zabbix_config_name: str, hosts_data: dict,
                    monitor_data: dict, filter_ids: list,
                    include_device_overview: bool, include_alert_data: bool,
                    zabbix_service, db: Session) -> tuple:
    """
    生成Excel报告
    返回 (文件路径, 文件名)
    """
    wb = Workbook()
    wb.remove(wb.active)

    # 设备概览页
    if include_device_overview:
        ws = wb.create_sheet("设备概览")
        ws.merge_cells('A1:C1')
        cell = ws.cell(row=1, column=1, value="设备概览")
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=3, column=1, value="设备总量").font = Font(bold=True, size=14)
        ws.cell(row=3, column=2, value=hosts_data["total"])
        ws.cell(row=4, column=1, value="在线设备量").font = Font(bold=True, size=14)
        ws.cell(row=4, column=2, value=len(hosts_data["online"]))
        ws.cell(row=5, column=1, value="离线设备量").font = Font(bold=True, size=14)
        ws.cell(row=5, column=2, value=len(hosts_data["offline"]))

        # 离线设备列表
        if hosts_data["offline"]:
            ws.cell(row=7, column=1, value="离线设备列表").font = Font(bold=True, size=14)
            headers = ["设备名", "IP地址", "所属群组"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=8, column=col, value=header).font = Font(bold=True)
            for idx, host in enumerate(hosts_data["offline"], 9):
                ws.cell(row=idx, column=1, value=host["name"]).font = Font(color="FF0000")
                ws.cell(row=idx, column=2, value=host["ip"]).font = Font(color="FF0000")
                ws.cell(row=idx, column=3, value=host["groups"]).font = Font(color="FF0000")

    # 为每个筛选配置创建工作表
    for filter_id in filter_ids:
        filter_obj = db.query(MonitorFilter).filter(MonitorFilter.id == filter_id).first()
        if not filter_obj:
            continue

        filter_data = monitor_data.get(filter_obj.name, [])
        if not filter_data:
            continue

        # 工作表名称最长31字符
        sheet_name = filter_obj.name[:31]
        ws = wb.create_sheet(sheet_name)

        # 标题
        ws.merge_cells('A1:H1')
        cell = ws.cell(row=1, column=1, value=filter_obj.name)
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # 表头
        headers = ["所属群组", "设备名", "IP地址", "监控项", "当前值", "最大值", "最小值", "平均值"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=header).font = Font(bold=True)

        # 数据
        for row_idx, row_data in enumerate(filter_data, 3):
            ws.cell(row=row_idx, column=1, value=row_data.get("groups", ""))
            ws.cell(row=row_idx, column=2, value=row_data.get("hostname", ""))
            ws.cell(row=row_idx, column=3, value=row_data.get("ip", ""))
            ws.cell(row=row_idx, column=4, value=row_data.get("item_name", ""))
            ws.cell(row=row_idx, column=5, value=row_data.get("current", 0))
            ws.cell(row=row_idx, column=6, value=row_data.get("max", 0))
            ws.cell(row=row_idx, column=7, value=row_data.get("min", 0))
            ws.cell(row=row_idx, column=8, value=row_data.get("avg", 0))

    # 添加告警数据工作表
    if include_alert_data:
        try:
            alerts = zabbix_service.get_alerts()
            if alerts:
                ws = wb.create_sheet("告警信息")

                # 标题
                ws.merge_cells('A1:I1')
                cell = ws.cell(row=1, column=1, value="告警信息")
                cell.font = Font(bold=True, size=16)
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # 表头
                headers = ["告警级别", "主机群组", "主机名", "主机IP", "告警信息",
                          "发生时间", "恢复时间", "持续时间", "状态"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=2, column=col, value=header).font = Font(bold=True)

                # 严重程度映射
                severity_map = {
                    0: "未分类",
                    1: "信息",
                    2: "警告",
                    3: "一般严重",
                    4: "严重",
                    5: "灾难"
                }

                # 格式化时间戳
                def format_timestamp(ts):
                    if ts and ts > 0:
                        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
                    return ""

                # 格式化持续时间
                def format_duration(seconds):
                    if not seconds or seconds <= 0:
                        return ""
                    days = seconds // 86400
                    hours = (seconds % 86400) // 3600
                    minutes = (seconds % 3600) // 60
                    secs = seconds % 60

                    parts = []
                    if days > 0:
                        parts.append(f"{days}天")
                    if hours > 0:
                        parts.append(f"{hours}小时")
                    if minutes > 0:
                        parts.append(f"{minutes}分钟")
                    if secs > 0 or not parts:
                        parts.append(f"{secs}秒")
                    return "".join(parts)

                # 数据
                for idx, alert in enumerate(alerts, 1):
                    row_idx = idx + 2
                    ws.cell(row=row_idx, column=1, value=severity_map.get(alert.get("severity", 0), "未知"))
                    ws.cell(row=row_idx, column=2, value=alert.get("host_groups", ""))
                    ws.cell(row=row_idx, column=3, value=alert.get("host_name", ""))
                    ws.cell(row=row_idx, column=4, value=alert.get("host_ip", ""))
                    ws.cell(row=row_idx, column=5, value=alert.get("name", ""))
                    ws.cell(row=row_idx, column=6, value=format_timestamp(alert.get("clock")))
                    ws.cell(row=row_idx, column=7, value=format_timestamp(alert.get("r_clock")) if alert.get("recovered") else "")
                    ws.cell(row=row_idx, column=8, value=format_duration(alert.get("duration", 0)))
                    ws.cell(row=row_idx, column=9, value="已恢复" if alert.get("recovered") else "告警中")
        except Exception as e:
            logger.error(f"生成告警数据工作表失败: {str(e)}")

    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"{zabbix_config_name}_{timestamp}.xlsx"
    filepath = os.path.join(EXPORT_DIR, filename)
    wb.save(filepath)

    return filepath, filename


class SchedulerService:
    """定时任务调度服务"""

    def __init__(self):
        self.scheduler = None
        self._initialized = False

    def init_scheduler(self):
        """初始化调度器"""
        if self._initialized:
            return

        # 使用内存存储任务
        jobstores = {
            'default': MemoryJobStore()
        }
        executors = {
            'default': ThreadPoolExecutor(5)
        }

        self.scheduler = BackgroundScheduler(
            jobstores=jobstores,
            executors=executors,
            timezone='Asia/Shanghai'
        )
        self.scheduler.start()
        self._initialized = True
        logger.info("调度器已启动")

    def shutdown(self):
        """关闭调度器"""
        if self.scheduler:
            self.scheduler.shutdown()
            self._initialized = False
            logger.info("调度器已关闭")

    def add_task(self, task: ScheduleTask):
        """添加任务"""
        if not self._initialized:
            self.init_scheduler()

        job_id = f"task_{task.id}"

        # 解析cron表达式
        parts = task.cron_expression.split()
        if len(parts) == 5:
            minute, hour, day, month, day_of_week = parts
        else:
            raise ValueError(f"无效的cron表达式: {task.cron_expression}")

        trigger = CronTrigger(
            minute=minute,
            hour=hour,
            day=day,
            month=month,
            day_of_week=day_of_week
        )

        self.scheduler.add_job(
            execute_task,  # 使用独立函数
            trigger=trigger,
            id=job_id,
            args=[task.id],
            replace_existing=True
        )

        logger.info(f"任务已添加: {task.name} (ID: {task.id})")

    def remove_task(self, task_id: int):
        """移除任务"""
        if not self._initialized:
            return

        job_id = f"task_{task_id}"
        try:
            self.scheduler.remove_job(job_id)
            logger.info(f"任务已移除: {job_id}")
        except Exception as e:
            logger.warning(f"移除任务失败: {e}")

    def update_task(self, task: ScheduleTask):
        """更新任务"""
        if task.is_active:
            self.remove_task(task.id)
            self.add_task(task)
        else:
            self.remove_task(task.id)

    def run_task_now(self, task_id: int):
        """立即执行任务"""
        execute_task(task_id)

    def load_all_tasks(self):
        """加载所有活跃任务"""
        db = SessionLocal()
        try:
            tasks = db.query(ScheduleTask).filter(ScheduleTask.is_active == True).all()
            for task in tasks:
                try:
                    self.add_task(task)
                except Exception as e:
                    logger.error(f"加载任务失败: {task.name} - {e}")
        finally:
            db.close()


# 全局实例
scheduler_service = SchedulerService()
